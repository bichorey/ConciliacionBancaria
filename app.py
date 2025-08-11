# -*- coding: utf-8 -*-

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from reconciliacion import (
    conciliacion_mvp,
    is_previous_result,
    extract_mayor_from_previous,
    merge_with_previous,
)

# --- Página ---
st.set_page_config(page_title="Conciliación Bancaria", layout="wide")
st.title("Conciliación Bancaria")

# --- IO helpers ---
@st.cache_data(show_spinner=False)
def _read_any(file):
    name = file.name.lower()
    if name.endswith((".xlsx", ".xls")):
        return pd.read_excel(file)
    else:
        # Infiero separador para CSV
        try:
            return pd.read_csv(file, sep=None, engine="python")  # autodetecta
        except Exception:
            file.seek(0)
            return pd.read_csv(file, sep=";", engine="python")

# --- Post-procesamiento de columnas ---
def _postprocess_detalle(detalle: pd.DataFrame) -> pd.DataFrame:
    # 0) Remover columnas duplicadas manteniendo la primera
    if detalle.columns.duplicated().any():
        # Log opcional: ver cuáles fueron duplicadas
        dupes = detalle.columns[detalle.columns.duplicated()].tolist()
        if dupes:
            st.info(f"Eliminando columnas duplicadas: {sorted(set(dupes))}")
        detalle = detalle.loc[:, ~detalle.columns.duplicated()].copy()

    # 1) Consolidar Importe_norm única
    if "Importe_norm" not in detalle.columns:
        if "Importe_norm_MAYOR" in detalle.columns or "Importe_norm_BANCO" in detalle.columns:
            detalle["Importe_norm"] = detalle.get("Importe_norm_MAYOR", pd.Series(dtype=float)).fillna(
                detalle.get("Importe_norm_BANCO", pd.Series(dtype=float))
            )
        else:
            detalle["Importe_norm"] = pd.NA

    # 2) Renombres exactos que vimos aparecer
    rename_for_spec = {
        "Nro. Comp_MAYOR": "Nro.Comp_MAYOR",
        "Razon,Social_MAYOR": "Razon Social_MAYOR",
    }
    detalle.rename(columns={k: v for k, v in rename_for_spec.items() if k in detalle.columns}, inplace=True)

    # 3) Orden exacto de columnas solicitado (completando faltantes como NA)
    desired_order = [
        "Código_MAYOR", "Cuenta_MAYOR", "Fecha_MAYOR", "Tipo_MAYOR", "Nro.Comp_MAYOR", "Subcuenta_MAYOR",
        "Detalle_MAYOR", "CUIT_MAYOR", "Razon Social_MAYOR", "Débito_MAYOR", "Crédito_MAYOR", "Saldo_MAYOR", "Importe_MAYOR",
        "NUM_BANCO", "FECHA_BANCO", "COMBTE_BANCO", "DESCRIPCION_BANCO", "DEBITO_BANCO", "CREDITO_BANCO", "SALDO_BANCO", "IMPORTE_BANCO",
        "Fecha_norm_MAYOR", "Fecha_norm_BANCO", "Importe_norm", "estado", "regla", "diferencia_dias", "grupo_id",
    ]

    # Asegurar que no haya duplicados en desired_order (por si alguien lo edita)
    desired_order = list(dict.fromkeys(desired_order))

    # Crear faltantes
    for col in desired_order:
        if col not in detalle.columns:
            detalle[col] = pd.NA

    # Antes del reindex, volver a chequear que no se hayan generado duplicados por renombres
    if detalle.columns.duplicated().any():
        detalle = detalle.loc[:, ~detalle.columns.duplicated()].copy()

    # Reindex seguro (ya sin duplicadas)
    detalle = detalle.reindex(columns=desired_order)

    return detalle

# --- Descarga Excel con formato ---
def _download_excel_button(detalle: pd.DataFrame, resumen: pd.DataFrame):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        detalle.to_excel(writer, index=False, sheet_name="Detalle")
        resumen.to_excel(writer, index=False, sheet_name="Resumen")

        wb = writer.book
        ws = wb["Detalle"]

        # Congelar encabezado (fila 1) y aplicar autofiltro
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Colores por estado
        fill_mayor = PatternFill(start_color="FFFDE9D9", end_color="FFFDE9D9", fill_type="solid")   # rojo suave
        fill_banco = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")   # verde suave

        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_map = {name: idx + 1 for idx, name in enumerate(header)}  # 1-based

        estado_col_idx = col_map.get("estado")
        fecha_m_col_idx = col_map.get("Fecha_norm_MAYOR")
        fecha_b_col_idx = col_map.get("Fecha_norm_BANCO")
        importe_norm_idx = col_map.get("Importe_norm")

        # Formatos
        date_fmt = "DD/MM/YY"
        currency_fmt = "$ #,##0.00"

        # Recorrer filas de datos
        for r in range(2, ws.max_row + 1):
            # Colorear por estado
            if estado_col_idx:
                val = ws.cell(row=r, column=estado_col_idx).value
                if val == "Solo en Mayor":
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = fill_mayor
                elif val == "Solo en Banco":
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = fill_banco
            # Formato de fechas
            if fecha_m_col_idx:
                ws.cell(row=r, column=fecha_m_col_idx).number_format = date_fmt
            if fecha_b_col_idx:
                ws.cell(row=r, column=fecha_b_col_idx).number_format = date_fmt
            # Formato moneda
            if importe_norm_idx:
                ws.cell(row=r, column=importe_norm_idx).number_format = currency_fmt

        # Autoajustar ancho de columnas según contenido
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[col_letter]:
                try:
                    val = "" if cell.value is None else str(cell.value)
                except Exception:
                    val = ""
                max_len = max(max_len, len(val))
            # Ajuste con límites razonables
            ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 60))

    output.seek(0)
    st.download_button(
        label="Descargar Excel (Detalle + Resumen)",
        data=output,
        file_name="Conciliacion_Bancaria.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# --- UI lateral ---
with st.sidebar:
    st.header("Parámetros")
    tolerancia_dias = st.number_input("Tolerancia de días", min_value=0, max_value=30, value=2, step=1)
    max_items_grupo = st.number_input("Máx. items por grupo (many-to-one)", min_value=1, max_value=10, value=4, step=1)
    direccion = st.selectbox("Dirección agrupación", ["MAYOR→BANCO", "BANCO→MAYOR"])
    st.caption("One-to-one siempre activo. Many-to-one sólo cuando máx. items > 1.")

# --- Carga de archivos ---
col1, col2 = st.columns(2)
with col1:
    mayor_file = st.file_uploader("Subí Mayor (xlsx o csv)", type=["xlsx", "xls", "csv"], key="mayor_up")
with col2:
    banco_file = st.file_uploader("Subí Banco (xlsx o csv)", type=["xlsx", "xls", "csv"], key="banco_up")

procesar = st.button("Conciliar")

# --- Lógica principal ---
if procesar:
    if not mayor_file or not banco_file:
        st.error("Subí ambos archivos (Mayor y Banco).")
        st.stop()

    try:
        df_mayor = _read_any(mayor_file)
        df_banco = _read_any(banco_file)
    except Exception as e:
        st.exception(e)
        st.stop()

    try:
        # Soporta archivo de resultado previo del Mayor:
        # - procesa sólo "Solo en Mayor"
        # - luego mergea para incluir TODAS las filas en el Excel final
        if is_previous_result(df_mayor):
            prev_detalle = df_mayor.copy()
            mayor_filtrado_prev = df_mayor[df_mayor["estado"] == "Solo en Mayor"].copy()
            df_mayor_base = extract_mayor_from_previous(mayor_filtrado_prev)

            detalle_nuevo, resumen = conciliacion_mvp(
                df_mayor_in=df_mayor_base,
                df_banco_in=df_banco,
                tolerancia_dias=int(tolerancia_dias),
                max_items_grupo=int(max_items_grupo),
                direccion=direccion,
            )
            detalle = merge_with_previous(prev_detalle, detalle_nuevo)
        else:
            detalle, resumen = conciliacion_mvp(
                df_mayor_in=df_mayor,
                df_banco_in=df_banco,
                tolerancia_dias=int(tolerancia_dias),
                max_items_grupo=int(max_items_grupo),
                direccion=direccion,
            )
    except Exception as e:
        st.error("Error durante la conciliación.")
        st.exception(e)
        st.stop()

    if detalle.empty:
        st.warning("No se generaron resultados.")
        st.stop()

    detalle = _postprocess_detalle(detalle)

    st.subheader("Resumen")
    st.dataframe(resumen, use_container_width=True)

    st.subheader("Detalle")
    st.dataframe(detalle, use_container_width=True)

    _download_excel_button(detalle, resumen)
else:
    st.info("Subí los archivos y hacé clic en Conciliar.")
    